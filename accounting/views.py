# -*- coding: utf-8 -*-
"""
SCMD Pro
------------------------------
Copyright (c) 2026 SCMD.co.ltd. All Rights Reserved.

File: accounting/views.py
Author: Mr. Anh
Created Date: 2025-12-04
Updated Date: 2026-06-03
Version: v3.5.0
Description: Views xử lý logic Kế toán & Lương.
             UPDATED: Bổ sung Data cho Dashboard KTT (KPIs, Charts).
"""

import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from main.dashboard_cta import admin_url_if_permitted, reverse_or_none
from django.contrib.auth.decorators import login_required
from main.dashboard_router import dashboard_access_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied, ValidationError
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from datetime import timedelta
import json
from typing import Any, cast, TYPE_CHECKING
from django.core.cache import cache
from .cache_utils import ACCOUNTING_DASHBOARD_UI_VERSION, build_accounting_dashboard_cache_key
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from .models import BangLuongThang, ChiTietLuong
from .models_soquy import SoQuy
from inventory.models import PhieuXuat
from .services.payroll import PayrollService
from accounting.application.reports_use_cases import DeductionAuditUseCase
from rolepermissions.checkers import has_role
from django.conf import settings

from main.audit_utils import export_audit_log

if TYPE_CHECKING:
    from core.managers import TenantAwareManager

logger = logging.getLogger(__name__)

EXPORT_ALLOWED_ROLES = ['ban_giam_doc', 'ke_toan']


def _enforce_export_access(request):
    if request.user.is_superuser or has_role(request.user, EXPORT_ALLOWED_ROLES):
        return
    raise PermissionDenied("Bạn không có quyền xuất dữ liệu lương nhạy cảm.")
# --- WEB ADMIN VIEWS (KTT/KTV) ---

@dashboard_access_required("accounting:dashboard")
def dashboard_accounting(request):
    """Accounting Workbench: read-only dashboard context for accounting users."""
    def _safe_reverse(viewname, *, args=None):
        return reverse_or_none(viewname, args=args)

    def _admin_url(viewname, permission_codename, *, args=None):
        return admin_url_if_permitted(
            request.user,
            viewname,
            permission_codename,
            args=args,
        )

    now = timezone.now()
    today = timezone.localdate()
    period = request.GET.get('period', 'month')
    if period == 'today':
        date_from = today
        period_label = 'Hôm nay'
    elif period == '7d':
        date_from = today - timedelta(days=6)
        period_label = '7 ngày gần nhất'
    else:
        date_from = today.replace(day=1)
        period = 'month'
        period_label = 'Tháng này'

    # --- CACHE LAYER (Rule 14) ---
    cache_key = build_accounting_dashboard_cache_key(
        request.user.id,
        period,
        ui_version=ACCOUNTING_DASHBOARD_UI_VERSION,
    )
    cached_response_data = cache.get(cache_key)
    if cached_response_data:
        logger.info(f"Dashboard accounting: Serving cached data for user {request.user.id}")
        return render(request, 'accounting/dashboard.html', cached_response_data)

    period_start = timezone.make_aware(
        timezone.datetime.combine(date_from, timezone.datetime.min.time()),
        timezone.get_current_timezone(),
    )
    first_day_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    tenant_id = str(getattr(settings, 'SCMD_ORGANIZATION_ID', ''))

    so_quy_mgr = cast("TenantAwareManager", SoQuy.objects)
    so_quy_period = so_quy_mgr.for_tenant(tenant_id).filter(ngay_lap__gte=period_start)
    so_quy_month = so_quy_mgr.for_tenant(tenant_id).filter(ngay_lap__gte=first_day_month)

    # Runtime v3: collapse dashboard totals into filtered aggregates instead of
    # issuing separate count/sum queries for each tile. This preserves the
    # single-organization tenant scope and avoids N+1-style query growth.
    month_totals = so_quy_month.aggregate(
        tong_thu=Sum('so_tien', filter=Q(trang_thai='DA_DUYET', loai_phieu='THU')),
        tong_chi=Sum('so_tien', filter=Q(trang_thai='DA_DUYET', loai_phieu='CHI')),
    )
    tong_thu = month_totals['tong_thu'] or 0
    tong_chi = month_totals['tong_chi'] or 0
    so_du_tam_tinh = tong_thu - tong_chi

    period_totals = so_quy_period.aggregate(
        period_thu=Sum('so_tien', filter=Q(trang_thai='DA_DUYET', loai_phieu='THU')),
        period_chi=Sum('so_tien', filter=Q(trang_thai='DA_DUYET', loai_phieu='CHI')),
    )
    period_thu = period_totals['period_thu'] or 0
    period_chi = period_totals['period_chi'] or 0
    period_balance = period_thu - period_chi

    pending_vouchers_qs = (
        so_quy_mgr.for_tenant(tenant_id)
        .filter(trang_thai='CHO_DUYET')
        .select_related('nhan_vien', 'hop_dong', 'nguoi_lap', 'nguoi_duyet')
        .order_by('-ngay_lap')
    )
    pending_voucher_stats = pending_vouchers_qs.aggregate(
        count=Count('id'),
        total=Sum('so_tien'),
    )
    phieu_cho_duyet = pending_voucher_stats['count'] or 0
    pending_voucher_total = pending_voucher_stats['total'] or 0
    pending_vouchers_list = list(pending_vouchers_qs[:6])

    # Rule 9: Cưỡng chế organization scope thông qua SSOT (settings.SCMD_ORGANIZATION_ID)
    # Tuyệt đối không fallback về .all() để bảo vệ ranh giới dữ liệu.
    payroll_mgr = cast("TenantAwareManager", BangLuongThang.objects)
    payroll_qs = payroll_mgr.for_tenant(tenant_id)
    payroll_attention_qs = payroll_qs.filter(
        trang_thai__in=[
            BangLuongThang.TrangThai.CALCULATED,
            BangLuongThang.TrangThai.REVIEWED,
        ]
    ).order_by('-nam', '-thang')
    payroll_attention_stats = payroll_attention_qs.aggregate(
        count=Count('id'),
        total=Sum('tong_chi_tra'),
    )
    luong_cho_phat_hanh = payroll_attention_stats['count'] or 0
    payroll_pending_total = payroll_attention_stats['total'] or 0
    payroll_attention_list = list(payroll_attention_qs[:6])
    current_payroll_attention = [
        payroll
        for payroll in payroll_attention_list
        if payroll.nam == now.year and payroll.thang == now.month
    ]
    payroll_work_list = (current_payroll_attention or payroll_attention_list)[:3]

    bang_luongs = list(payroll_qs.order_by('-nam', '-thang')[:12])
    can_view_bangluong = request.user.has_perm('accounting.view_bangluongthang')
    can_change_bangluong = request.user.has_perm('accounting.change_bangluongthang')
    for bl in bang_luongs:
        bl.detail_url = _safe_reverse('accounting:bang_luong_detail', args=[bl.pk]) if can_view_bangluong else None
    latest_payroll = bang_luongs[0] if bang_luongs else None
    payroll_status_stats = payroll_qs.aggregate(
        paid=Count('id', filter=Q(trang_thai=BangLuongThang.TrangThai.PAID)),
        locked=Count('id', filter=Q(trang_thai=BangLuongThang.TrangThai.LOCKED)),
    )
    paid_payroll_count = payroll_status_stats['paid'] or 0
    locked_payroll_count = payroll_status_stats['locked'] or 0

    recent_vouchers = list(
        so_quy_mgr.for_tenant(tenant_id)
        .select_related('nhan_vien', 'hop_dong', 'nguoi_lap')
        .order_by('-ngay_lap')[:8]
    )

    category_rows = list(
        so_quy_period
        .values('hang_muc')
        .annotate(total=Sum('so_tien'), count=Count('id'))
        .order_by('-total')[:6]
    )
    category_label_map = dict(SoQuy.HANG_MUC)
    max_category_total = max([row['total'] or 0 for row in category_rows] or [1]) or 1
    for row in category_rows:
        row['label'] = category_label_map.get(row['hang_muc'], row['hang_muc'])
        row['percent'] = round(((row['total'] or 0) / max_category_total) * 100)

    # 6 tháng gần nhất - giữ biến cũ cho chart canvas nếu template khác còn dùng.
    chart_labels = []
    data_thu = []
    data_chi = []
    cashflow_points = []

    def _month_start(base_dt, months_ago=0):
        month = base_dt.month - months_ago
        year = base_dt.year
        while month <= 0:
            month += 12
            year -= 1
        while month > 12:
            month -= 12
            year += 1
        return base_dt.replace(year=year, month=month, day=1, hour=0, minute=0, second=0, microsecond=0)

    month_starts = [_month_start(now, i) for i in range(5, -1, -1)]
    cashflow_start = month_starts[0]
    cashflow_end = _month_start(now, -1)
    cashflow_rows = (
        so_quy_mgr.for_tenant(tenant_id)
        .filter(trang_thai='DA_DUYET', ngay_lap__gte=cashflow_start, ngay_lap__lt=cashflow_end)
        .annotate(month_bucket=TruncMonth('ngay_lap'))
        .values('month_bucket')
        .annotate(
            thu=Sum('so_tien', filter=Q(loai_phieu='THU')),
            chi=Sum('so_tien', filter=Q(loai_phieu='CHI')),
        )
    )
    cashflow_by_month = {}
    for row in cashflow_rows:
        bucket = row.get('month_bucket')
        if not bucket:
            continue
        bucket_date = bucket.date() if hasattr(bucket, 'date') else bucket
        cashflow_by_month[bucket_date] = row

    for month_start in month_starts:
        label = f'T{month_start.month}'
        row = cashflow_by_month.get(month_start.date(), {})
        thu = row.get('thu') or 0
        chi = row.get('chi') or 0
        chart_labels.append(label)
        data_thu.append(float(thu))
        data_chi.append(float(chi))
        cashflow_points.append({'label': label, 'thu': thu, 'chi': chi, 'balance': thu - chi})

    max_cashflow_value = max([max(point['thu'], point['chi']) for point in cashflow_points] or [1]) or 1
    cashflow_has_activity = any(point['thu'] or point['chi'] for point in cashflow_points)
    for point in cashflow_points:
        point['thu_percent'] = round((point['thu'] / max_cashflow_value) * 100) if max_cashflow_value else 0
        point['chi_percent'] = round((point['chi'] / max_cashflow_value) * 100) if max_cashflow_value else 0

    payslip_mgr = cast("TenantAwareManager", ChiTietLuong.objects)
    deduction_scope_qs = payslip_mgr.for_tenant(tenant_id).filter(
        bang_luong__thang=now.month,
        bang_luong__nam=now.year,
        bang_luong__trang_thai__in=[
            BangLuongThang.TrangThai.CALCULATED,
            BangLuongThang.TrangThai.REVIEWED,
            BangLuongThang.TrangThai.LOCKED,
        ],
    )
    deduction_stats = deduction_scope_qs.aggregate(
        count=Count('id'),
        ung_luong=Sum('ung_luong'),
        phat_vi_pham=Sum('phat_vi_pham'),
        tien_dong_phuc=Sum('tien_dong_phuc'),
        tien_den_bu=Sum('tien_den_bu'),
        bao_hiem=Sum('bao_hiem'),
        phi_cong_doan=Sum('phi_cong_doan'),
    )
    if not deduction_stats.get('count'):
        deduction_scope_qs = payslip_mgr.for_tenant(tenant_id).filter(
            bang_luong__in=payroll_attention_qs[:3]
        )
        deduction_stats = deduction_scope_qs.aggregate(
            count=Count('id'),
            ung_luong=Sum('ung_luong'),
            phat_vi_pham=Sum('phat_vi_pham'),
            tien_dong_phuc=Sum('tien_dong_phuc'),
            tien_den_bu=Sum('tien_den_bu'),
            bao_hiem=Sum('bao_hiem'),
            phi_cong_doan=Sum('phi_cong_doan'),
        )
        deduction_scope_label = '3 kỳ lương đang xử lý gần nhất'
    else:
        deduction_scope_label = 'Kỳ hiện tại và kỳ đang xử lý'
    deduction_total = sum(
        deduction_stats.get(key) or 0
        for key in (
            'ung_luong',
            'phat_vi_pham',
            'tien_dong_phuc',
            'tien_den_bu',
            'bao_hiem',
            'phi_cong_doan',
        )
    )
    deduction_record_count = deduction_stats.get('count') or 0

    can_add_soquy = request.user.has_perm('accounting.add_soquy')
    can_view_soquy = request.user.has_perm('accounting.view_soquy')
    can_change_soquy = request.user.has_perm('accounting.change_soquy')
    can_view_cauhinhluong = request.user.has_perm('accounting.view_cauhinhluong')
    can_view_chitietluong = request.user.has_perm('accounting.view_chitietluong')
    can_run_payroll = (request.user.is_superuser or has_role(request.user, ["ban_giam_doc", "ke_toan"])) and can_change_bangluong

    urls = {
        'add_voucher': _admin_url('admin:accounting_soquy_add', 'accounting.add_soquy') if can_add_soquy else None,
        'vouchers': _admin_url('admin:accounting_soquy_changelist', 'accounting.view_soquy') if can_view_soquy else None,
        'payrolls': _admin_url('admin:accounting_bangluongthang_changelist', 'accounting.view_bangluongthang') if can_view_bangluong else None,
        'salary_config': _admin_url('admin:accounting_cauhinhluong_changelist', 'accounting.view_cauhinhluong') if can_view_cauhinhluong else None,
        'payslips': _admin_url('admin:accounting_chitietluong_changelist', 'accounting.view_chitietluong') if can_view_chitietluong else None,
        'calculate_payroll': _safe_reverse('accounting:tinh_luong') if can_run_payroll else None,
    }

    def _priority_badge(tone):
        return {
            'danger': 'Cao',
            'warning': 'Vừa',
            'info': 'Thấp',
        }.get(tone, 'Thấp')

    def _payroll_status(payroll):
        if payroll.trang_thai == BangLuongThang.TrangThai.REVIEWED:
            return 'Chờ duyệt/phát hành'
        if payroll.trang_thai == BangLuongThang.TrangThai.CALCULATED:
            return 'Cần đối soát'
        return payroll.get_trang_thai_display()

    action_items = []
    for payroll in payroll_work_list:
        payroll_url = _safe_reverse('accounting:bang_luong_detail', args=[payroll.pk]) if can_view_bangluong else None
        if payroll_url:
            tone = 'danger' if payroll.trang_thai == BangLuongThang.TrangThai.REVIEWED else 'warning'
            action_items.append({
                'priority': _priority_badge(tone),
                'type': 'Đối soát kỳ lương' if payroll.trang_thai == BangLuongThang.TrangThai.CALCULATED else 'Mở kỳ lương chờ phát hành',
                'title': f'Tháng {payroll.thang}/{payroll.nam}',
                'subject': payroll.ten_bang_luong,
                'amount': payroll.tong_chi_tra,
                'status': _payroll_status(payroll),
                'due_label': payroll.ngay_chot_cong.strftime('%d/%m') if payroll.ngay_chot_cong else '—',
                'url': payroll_url,
                'cta': 'Đối soát' if payroll.trang_thai == BangLuongThang.TrangThai.CALCULATED else 'Mở kỳ lương',
                'tone': tone,
            })

    for voucher in pending_vouchers_list[:4]:
        voucher_url = _admin_url('admin:accounting_soquy_change', 'accounting.change_soquy', args=[voucher.pk]) if can_change_soquy else None
        if voucher_url:
            tone = 'danger' if voucher.loai_phieu == 'CHI' else 'warning'
            action_items.append({
                'priority': _priority_badge(tone),
                'type': 'Xử lý phiếu chi' if voucher.loai_phieu == 'CHI' else 'Xử lý phiếu thu',
                'title': voucher.ma_phieu,
                'subject': voucher.get_hang_muc_display(),
                'amount': voucher.so_tien,
                'status': voucher.get_trang_thai_display(),
                'due_label': voucher.ngay_lap.strftime('%d/%m %H:%M') if voucher.ngay_lap else '—',
                'url': voucher_url,
                'cta': 'Mở phiếu',
                'tone': tone,
            })

    if urls['vouchers'] and (period_balance or period_thu or period_chi):
        action_items.append({
            'priority': 'Vừa' if period_balance < 0 else 'Thấp',
            'type': 'Rà soát sổ quỹ',
            'title': period_label,
            'subject': 'Thu - chi đã duyệt',
            'amount': period_balance,
            'status': 'Cần kiểm tra' if period_balance < 0 else 'Bình thường',
            'due_label': today.strftime('%d/%m'),
            'url': urls['vouchers'],
            'cta': 'Kiểm tra',
            'tone': 'warning' if period_balance < 0 else 'info',
        })

    tone_rank = {'danger': 0, 'warning': 1, 'info': 2}
    action_items = sorted(
        action_items,
        key=lambda item: (tone_rank.get(item['tone'], 3), item['due_label']),
    )[:8]

    metric_cards = [
        {
            'label': 'Kỳ lương cần xử lý',
            'value': luong_cho_phat_hanh,
            'status': 'Cảnh báo' if luong_cho_phat_hanh else 'Tốt',
            'cta': 'Xem',
            'tone': 'purple' if luong_cho_phat_hanh else 'slate',
            'url': urls['payrolls'],
        },
        {
            'label': 'Phiếu chờ duyệt',
            'value': phieu_cho_duyet,
            'status': 'Chờ duyệt' if phieu_cho_duyet else 'Tốt',
            'cta': 'Xem',
            'tone': 'amber' if phieu_cho_duyet else 'slate',
            'url': urls['vouchers'],
        },
        {
            'label': 'Số quỹ',
            'value': so_du_tam_tinh,
            'status': 'Cần kiểm tra' if so_du_tam_tinh < 0 else 'Bình thường',
            'cta': 'Mở',
            'tone': 'red' if so_du_tam_tinh < 0 else 'blue',
            'url': urls['vouchers'],
            'money': True,
        },
        {
            'label': 'Thu tháng này',
            'value': tong_thu,
            'status': 'Đã duyệt',
            'cta': 'Xem',
            'tone': 'green',
            'url': urls['vouchers'],
            'money': True,
        },
        {
            'label': 'Chi tháng này',
            'value': tong_chi,
            'status': 'Đã duyệt',
            'cta': 'Xem',
            'tone': 'red',
            'url': urls['vouchers'],
            'money': True,
        },
        {
            'label': 'Khấu trừ cần đối soát',
            'value': deduction_total,
            'status': f'{deduction_record_count} hồ sơ · {deduction_scope_label}' if deduction_record_count else deduction_scope_label,
            'cta': 'Đối soát',
            'tone': 'amber' if deduction_total else 'slate',
            'url': urls['payslips'],
            'money': True,
        },
    ]
    metric_cards = [card for card in metric_cards if card.get('url')]

    utility_cards = []
    if cashflow_has_activity:
        utility_cards.append({
            'kind': 'cashflow',
            'title': 'Dòng tiền 6 tháng',
        })

    accounting_priority_cards = []
    recent_payroll_rows = bang_luongs[:6]
    recent_voucher_rows = recent_vouchers[:6]
    context = {
        # Biến cũ giữ tương thích.
        'kpi_thu': tong_thu,
        'kpi_chi': tong_chi,
        'kpi_so_du': so_du_tam_tinh,
        'pending_phieu': phieu_cho_duyet,
        'pending_luong': luong_cho_phat_hanh,
        'chart_labels': json.dumps(chart_labels),
        'data_thu': json.dumps(data_thu),
        'data_chi': json.dumps(data_chi),
        'bang_luongs': bang_luongs,
        'today': now,

        # Biến mới cho dashboard kế toán.
        'period': period,
        'period_label': period_label,
        'date_from': date_from,
        'period_thu': period_thu,
        'period_chi': period_chi,
        'period_balance': period_balance,
        'pending_voucher_total': pending_voucher_total,
        'payroll_pending_total': payroll_pending_total,
        'latest_payroll': latest_payroll,
        'paid_payroll_count': paid_payroll_count,
        'locked_payroll_count': locked_payroll_count,
        'recent_vouchers': recent_vouchers,
        'pending_vouchers': pending_vouchers_list,
        'payroll_attention': payroll_attention_list,
        'category_rows': category_rows,
        'cashflow_points': cashflow_points,
        'cashflow_has_activity': cashflow_has_activity,
        'max_cashflow_value': max_cashflow_value,
        'action_items': action_items,
        'metric_cards': metric_cards,
        'accounting_priority_cards': accounting_priority_cards,
        'deduction_total': deduction_total,
        'deduction_record_count': deduction_record_count,
        'deduction_scope_label': deduction_scope_label,
        'recent_payroll_rows': recent_payroll_rows,
        'recent_voucher_rows': recent_voucher_rows,
        'utility_cards': utility_cards,
        'can_run_payroll': can_run_payroll,
        'urls': urls,
        'accounting_dashboard_cache_version': ACCOUNTING_DASHBOARD_UI_VERSION,
    }

    # Lưu vào cache trong 1 giờ (Sẽ bị invalidated tự động khi có signal thay đổi dữ liệu)
    cache.set(cache_key, context, 3600)
    return render(request, 'accounting/dashboard.html', context)

@login_required
def tinh_luong_view(request):
    if request.method == "POST":
        try:
            thang = int(request.POST.get('thang'))
            nam = int(request.POST.get('nam'))
            success, msg = PayrollService.tinh_luong_thang(thang, nam)
            if success: messages.success(request, msg)
            else: messages.error(request, msg)
        except Exception as e:
            messages.error(request, f"Lỗi: {str(e)}")
    return redirect('accounting:dashboard')

@login_required
def chi_tiet_bang_luong(request, pk):
    bl = get_object_or_404(BangLuongThang, pk=pk)
    chi_tiet = ChiTietLuong.objects.filter(bang_luong=bl).select_related('nhan_vien')
    return render(request, 'accounting/bang_luong_detail.html', {'bang_luong': bl, 'chi_tiet': chi_tiet})

@login_required
def chot_luong_view(request, pk):
    """
    View chốt bảng lương.
    Refactor: Chuyển logic nghiệp vụ vào LockPayrollUseCase để đảm bảo Audit/Permission (P1).
    """
    from accounting.application.payroll_lock_use_case import LockPayrollUseCase

    try:
        tenant_id = str(getattr(settings, 'SCMD_ORGANIZATION_ID', ''))
        payroll, changed = LockPayrollUseCase.execute(
            payroll_id=pk,
            actor_user=request.user,
            tenant_id=tenant_id,
            reason="Thực hiện chốt lương định kỳ qua Technical Console."
        )
        if changed:
            messages.success(request, f"Đã khóa bảng lương tháng {payroll.thang}/{payroll.nam} thành công.")
        else:
            messages.info(request, "Bảng lương này đã được khóa trước đó.")
    except (PermissionDenied, ValidationError) as e:
        messages.error(request, str(e))
    except Exception:
        logger.exception("Unexpected system error in chot_luong_view for payroll %s", pk)
        messages.error(request, "Có lỗi hệ thống xảy ra khi chốt lương. Vui lòng liên hệ quản trị.")

    return redirect('accounting:dashboard')

@login_required
def bao_cao_doi_soat_khau_tru(request, pk):
    """View hiển thị báo cáo đối soát khấu trừ lương vs Sổ quỹ"""
    tenant_id = str(getattr(settings, 'SCMD_ORGANIZATION_ID', ''))
    report_data = DeductionAuditUseCase.execute(pk, tenant_id)
    
    return render(request, 'accounting/reports/deduction_audit.html', {'report': report_data})

@login_required
@export_audit_log(
    module="accounting",
    model_name="BangLuongThang",
    note="Export Excel doi soat khau tru",
    object_id_resolver=lambda request, pk: pk,
    changes_resolver=lambda request, pk: {"bang_luong_id": pk},
)
def export_doi_soat_khau_tru_excel(request, pk):
    """Xuất file Excel báo cáo đối soát khấu trừ lương"""
    _enforce_export_access(request)
    tenant_id = str(getattr(settings, 'SCMD_ORGANIZATION_ID', ''))
    report = DeductionAuditUseCase.execute(pk, tenant_id)
    
    if report['status'] == 'error':
        messages.error(request, f"Lỗi xuất file: {report['message']}")
        return redirect('accounting:doi_soat_khau_tru', pk=pk)

    # Khởi tạo Workbook
    wb = Workbook()
    ws = wb.active # type: ignore
    if ws is None:
        return HttpResponse("Lỗi khởi tạo Excel.", status=500)
    ws.title = "Doi Soat Khau Tru"

    # Định dạng header
    headers = ['STT', 'Nhân viên', 'Mã NV', 'Khấu trừ Lương (A)', 'Thực chi Sổ quỹ (B)', 'Chênh lệch (A-B)']
    ws.append([f"BÁO CÁO ĐỐI SOÁT: {report['bang_luong']}"])
    ws.append([f"Kỳ lương: {report['ky_luong']}"])
    ws.append([]) # Dòng trống
    ws.append(headers)

    # Định dạng cảnh báo cho sai lệch (Màu đỏ)
    warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    warning_font = Font(color='9C0006', bold=True)

    # Định dạng Accounting (VND) cho Excel
    currency_format = '#,##0" ₫"'
    # Tối ưu độ rộng các cột để hiển thị số tiền không bị che khuất
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    # Thêm dữ liệu
    for i, item in enumerate(cast(list[dict[str, Any]], report['data']), 1):
        ws.append([
            i,
            item['nhan_vien'],
            item['ma_nv'],
            item['khau_tru_luong'],
            item['thuc_chi_so_quy'],
            item['chenh_lech']
        ])
        # Áp dụng định dạng tiền tệ cho các cột D, E, F (tương ứng index 4, 5, 6)
        for col_num in range(4, 7):
            ws.cell(row=ws.max_row, column=col_num).number_format = currency_format

        # Cảnh báo chênh lệch khác 0 (Cột F - Index 6)
        if item['chenh_lech'] != 0:
            cell = ws.cell(row=ws.max_row, column=6)
            cell.fill = warning_fill
            cell.font = warning_font

    # Thêm dòng tổng kết
    ws.append([])
    summary = cast(dict[str, Any], report.get('summary', {}))
    ws.append(['', '', 'TỔNG BIẾN ĐỘNG:', '', '', summary.get('total_variance', 0)])
    # Định dạng cho ô tổng chênh lệch ở cột F
    total_variance_cell = ws.cell(row=ws.max_row, column=6)
    total_variance_cell.number_format = currency_format
    
    # Nếu tổng biến động khác 0, cũng tô màu cảnh báo
    if summary.get('total_variance', 0) != 0:
        total_variance_cell.fill = warning_fill
        total_variance_cell.font = warning_font
        total_variance_cell.font = Font(color='9C0006', bold=True, size=12) # Nhấn mạnh thêm cho dòng tổng

    # --- BẢO VỆ DỮ LIỆU (SECURITY BY DESIGN) ---
    # Khóa Sheet và đặt mật khẩu để đảm bảo tính toàn vẹn, ngăn chặn chỉnh sửa trái phép
    ws.protection.sheet = True
    ws.protection.set_password(settings.EXCEL_EXPORT_PASSWORD)

    # Thiết lập Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"Doi_soat_luong_{report['ky_luong'].replace('/', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# --- MOBILE VIEWS (CHO BẢO VỆ) ---

@login_required
def mobile_phieu_luong_list(request):
    try: nv = request.user.nhan_vien
    except: return redirect('operations:mobile_dashboard')
    
    phieu_luongs = ChiTietLuong.objects.filter(
        nhan_vien=nv, 
        bang_luong__trang_thai__in=[
            BangLuongThang.TrangThai.LOCKED,
            BangLuongThang.TrangThai.PAID,
        ]
    ).select_related('bang_luong').order_by('-bang_luong__nam', '-bang_luong__thang')
    
    return render(request, 'accounting/mobile/phieu_luong_list.html', {'phieu_luongs': phieu_luongs})

@login_required
def mobile_phieu_luong_detail(request, pk):
    phieu = get_object_or_404(ChiTietLuong, pk=pk, nhan_vien=request.user.nhan_vien)
    return render(request, 'accounting/mobile/phieu_luong_detail.html', {'p': phieu})
