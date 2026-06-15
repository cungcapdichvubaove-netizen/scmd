# -*- coding: utf-8 -*-
<<<<<<< HEAD
"""Report export services for SCMD Pro.

This module generates PDF/Excel outputs from operational data.  All
user-facing Vietnamese strings in this file must remain valid UTF-8; do not
paste mojibake text into reports because exported documents are audit records.
"""

import csv
import io
from datetime import date

import openpyxl
from django.conf import settings
from django.core.exceptions import ValidationError
from django.http import HttpResponse
=======
"""
Security Command (SCMD) System
------------------------------
Copyright (c) 2025 SCMD.co.ltd. All Rights Reserved.

File: reports/services.py
Author: Mr. Anh
Created Date: 2025-12-10
Description: Service xá»­ lÃ½ xuáº¥t bÃ¡o cÃ¡o (PDF/Excel).
             - Sá»­ dá»¥ng WeasyPrint Ä‘á»ƒ táº¡o PDF tá»« HTML template.
             - Sá»­ dá»¥ng OpenPyXL Ä‘á»ƒ táº¡o bÃ¡o cÃ¡o Excel chuyÃªn nghiá»‡p.
"""

import io

import openpyxl
from django.conf import settings
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from weasyprint import HTML

<<<<<<< HEAD
from main.company_info import get_company_report_context

from operations.access_policies import IncidentVisibilityPolicy, ShiftVisibilityPolicy
from operations.models import BaoCaoSuCo, PhanCongCaTruc


# SSOT report style tokens, mapped from static/common/css/brand_system.css.
REPORT_BRAND_HEADER_FILL = "16233A"  # --scmd-navy-900
REPORT_BRAND_HEADER_TEXT = "FFFFFF"  # --scmd-on-brand
REPORT_BORDER_STYLE = "thin"
REPORT_ATTENDANCE_EXCEL_MAX_ROWS_DEFAULT = 10000

INCIDENT_CSV_HEADERS = [
    "Thời gian",
    "Mã",
    "Tiêu đề",
    "Mục tiêu",
    "Người báo",
    "Mức độ",
    "Trạng thái",
    "Thiệt hại",
]


class ReportService:
    @staticmethod
    def generate_incident_csv_response(incident_queryset, month, year):
        """Build incident CSV from an already scoped queryset.

        The caller is responsible for RBAC/export audit and for passing the
        queryset returned by GetIncidentReportUseCase with the request user. This
        keeps view code thin without weakening IncidentVisibilityPolicy.
        """
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = f'attachment; filename="SuCo_T{month}_{year}.csv"'
        writer = csv.writer(response)
        writer.writerow(INCIDENT_CSV_HEADERS)
        for sc in incident_queryset:
            writer.writerow([
                sc.created_at.strftime("%d/%m/%Y %H:%M"),
                sc.ma_su_co,
                sc.tieu_de,
                sc.muc_tieu.ten_muc_tieu if sc.muc_tieu else "",
                sc.nhan_vien_bao_cao.ho_ten if sc.nhan_vien_bao_cao else "",
                sc.get_muc_do_display(),
                sc.get_trang_thai_display(),
                sc.tong_thiet_hai,
            ])
        return response

    @staticmethod
    def generate_incident_pdf(incident_id, request=None, tenant_id=None, user=None):
        """Tạo file PDF cho một sự cố cụ thể."""
        try:
            scoped_tenant_id = tenant_id or getattr(settings, "SCMD_ORGANIZATION_ID", None)
            incident_qs = (
                IncidentVisibilityPolicy.visible_incidents(user)
                if user is not None
                else BaoCaoSuCo.objects.for_tenant(scoped_tenant_id)
            )
            incident = incident_qs.get(id=incident_id)

            context = {
                "incident": incident,
                "company_info": get_company_report_context(),
=======
from operations.models import BaoCaoSuCo, PhanCongCaTruc


class ReportService:
    @staticmethod
    def generate_incident_pdf(incident_id, request=None, tenant_id=None):
        """
        Táº¡o file PDF cho má»™t sá»± cá»‘ cá»¥ thá»ƒ.
        """
        try:
            scoped_tenant_id = tenant_id or getattr(settings, "SCMD_ORGANIZATION_ID", None)
            incident = BaoCaoSuCo.objects.for_tenant(scoped_tenant_id).get(id=incident_id)

            context = {
                "incident": incident,
                "company_info": {
                    "name": "CÃ”NG TY Dá»ŠCH Vá»¤ Báº¢O Vá»† SCMD",
                    "address": "123 ÄÆ°á»ng Sá»‘ 1, Quáº­n 1, TP.HCM",
                    "hotline": "1900 1234",
                },
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34
                "print_time": timezone.now(),
            }
            html_string = render_to_string("reports/print/incident_pdf.html", context, request=request)
            base_url = request.build_absolute_uri("/") if request else settings.BASE_DIR

            pdf_file = io.BytesIO()
            HTML(string=html_string, base_url=base_url).write_pdf(target=pdf_file)
            pdf_file.seek(0)
            return pdf_file, f"SuCo_{incident.ma_su_co}.pdf"

        except BaoCaoSuCo.DoesNotExist:
            return None, None

    @staticmethod
<<<<<<< HEAD
    def generate_attendance_excel(month, year, muc_tieu_id=None, tenant_id=None, user=None):
        """Xuất báo cáo chấm công tháng ra Excel.

        Performance guard:
        - filter the month through a sargable date range instead of
          ``__month``/``__year`` lookups;
        - keep the queryset scoped through ``ShiftVisibilityPolicy``;
        - stop oversized synchronous exports before tying up a web worker.
=======
    def generate_attendance_excel(month, year, muc_tieu_id=None, tenant_id=None):
        """
        Xuáº¥t bÃ¡o cÃ¡o cháº¥m cÃ´ng thÃ¡ng ra Excel.
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"ChamCong_T{month}_{year}"

<<<<<<< HEAD
        header_font = Font(bold=True, color=REPORT_BRAND_HEADER_TEXT)
        header_fill = PatternFill(start_color=REPORT_BRAND_HEADER_FILL, end_color=REPORT_BRAND_HEADER_FILL, fill_type="solid")
        border_style = Side(style=REPORT_BORDER_STYLE)
        full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        ws.merge_cells("A1:G1")
        ws["A1"] = f"BẢNG TỔNG HỢP CHẤM CÔNG - THÁNG {month}/{year}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["STT", "Ngày trực", "Nhân viên", "Mục tiêu", "Ca trực", "Giờ vào/ra", "Trạng thái"]
=======
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        border_style = Side(style="thin")
        full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Báº¢NG Tá»”NG Há»¢P CHáº¤M CÃ”NG - THÃNG {month}/{year}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["STT", "NgÃ y trá»±c", "NhÃ¢n viÃªn", "Má»¥c tiÃªu", "Ca trá»±c", "Giá» vÃ o/ra", "Tráº¡ng thÃ¡i"]
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34
        ws.append([])
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        scoped_tenant_id = tenant_id or getattr(settings, "SCMD_ORGANIZATION_ID", None)
<<<<<<< HEAD
        if month == 12:
            period_start = date(year, 12, 1)
            period_end = date(year + 1, 1, 1)
        else:
            period_start = date(year, month, 1)
            period_end = date(year, month + 1, 1)

        shift_qs = (
            ShiftVisibilityPolicy.visible_shifts(user)
            if user is not None
            else PhanCongCaTruc.objects.for_tenant(scoped_tenant_id)
        )
        queryset = (
            shift_qs
            .filter(ngay_truc__gte=period_start, ngay_truc__lt=period_end)
            .select_related("nhan_vien", "vi_tri_chot__muc_tieu", "ca_lam_viec", "chamcong")
=======
        queryset = (
            PhanCongCaTruc.objects.for_tenant(scoped_tenant_id)
            .filter(ngay_truc__month=month, ngay_truc__year=year)
            .select_related("nhan_vien", "vi_tri_chot__muc_tieu", "ca_lam_viec")
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34
        )

        if muc_tieu_id:
            queryset = queryset.filter(vi_tri_chot__muc_tieu_id=muc_tieu_id)

        queryset = queryset.order_by("ngay_truc", "nhan_vien__ho_ten")
<<<<<<< HEAD
        row_count = queryset.count()
        max_rows = int(
            getattr(
                settings,
                "SCMD_REPORT_ATTENDANCE_EXCEL_MAX_ROWS",
                REPORT_ATTENDANCE_EXCEL_MAX_ROWS_DEFAULT,
            )
        )
        if row_count > max_rows:
            raise ValidationError(
                f"Báo cáo có {row_count} dòng, vượt ngưỡng xuất đồng bộ {max_rows} dòng. "
                "Vui lòng lọc theo mục tiêu hoặc khoảng dữ liệu hẹp hơn."
            )

        for idx, pc in enumerate(queryset, 1):
            cc_info = "Chưa chấm"
            status = "Vắng"
=======

        for idx, pc in enumerate(queryset, 1):
            cc_info = "ChÆ°a cháº¥m"
            status = "Váº¯ng"
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34
            if hasattr(pc, "chamcong"):
                cc = pc.chamcong
                if cc.thoi_gian_check_in:
                    in_time = cc.thoi_gian_check_in.strftime("%H:%M")
                    out_time = cc.thoi_gian_check_out.strftime("%H:%M") if cc.thoi_gian_check_out else "--:--"
                    cc_info = f"{in_time} - {out_time}"
<<<<<<< HEAD
                    status = "Hoàn thành" if cc.thoi_gian_check_out else "Đang trực"
=======
                    status = "HoÃ n thÃ nh" if cc.thoi_gian_check_out else "Äang trá»±c"
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34

            ws.append(
                [
                    idx,
                    pc.ngay_truc.strftime("%d/%m/%Y"),
                    pc.nhan_vien.ho_ten,
                    pc.vi_tri_chot.muc_tieu.ten_muc_tieu,
                    pc.ca_lam_viec.ten_ca,
                    cc_info,
                    status,
                ]
            )

            for col in range(1, 8):
                ws.cell(row=3 + idx, column=col).border = full_border

        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["F"].width = 20

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

<<<<<<< HEAD
        return excel_file, f"BangCong_T{month}_{year}.xlsx", row_count
=======
        return excel_file, f"BangCong_T{month}_{year}.xlsx"
>>>>>>> 51661ed7e1165a088e9f7635fb9a4a3d23400f34
